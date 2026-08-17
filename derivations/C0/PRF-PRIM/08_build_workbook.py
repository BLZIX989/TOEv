"""
PRF-PRIM -- assemble all computed results into results/workbooks/C0_PRF_PRIM_Primitive_Reconciliation.xlsx

Run: python3 08_build_workbook.py
"""
import json, csv, os
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(__file__)
DOUT = os.path.join(HERE, 'output')
COUT = os.path.normpath(os.path.join(HERE, '..', '..', '..', 'calculations', 'C0', 'PRF-PRIM', 'output'))
GOUT = os.path.normpath(os.path.join(HERE, '..', '..', '..', 'graphs', 'C0', 'PRF-PRIM'))
WB_OUT = os.path.normpath(os.path.join(HERE, '..', '..', '..', 'results', 'workbooks'))

HEADER_FILL = PatternFill(start_color='7A1F2B', end_color='7A1F2B', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')

def load_json(path):
    with open(path) as f:
        return json.load(f)

def load_csv(path):
    with open(path, newline='') as f:
        return list(csv.reader(f))

def write_rows(ws, rows, start_row=1):
    r = start_row
    if not rows:
        ws.cell(row=r, column=1, value='(no data)')
        return
    header = rows[0]
    for c, val in enumerate(header, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for ri, row in enumerate(rows[1:], start=r + 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=ri, column=c, value=str(val) if not isinstance(val, (int, float, type(None))) else val)
    for c in range(1, len(header) + 1):
        col = get_column_letter(c)
        maxlen = max([len(str(header[c-1]))] + [len(str(row[c-1])) if c-1 < len(row) else 0 for row in rows[1:]])
        ws.column_dimensions[col].width = min(max(12, maxlen * 0.9), 70)
    ws.freeze_panes = 'A2'

def dict_to_rows(d, cols=("Key", "Value")):
    rows = [list(cols)]
    for k, v in d.items():
        rows.append([k, json.dumps(v) if isinstance(v, (dict, list)) else v])
    return rows

def nested_json_to_rows(d, id_col_name="Item"):
    """Flatten a {name: {field: value, ...}} dict into a table."""
    rows = None
    out = []
    for name, fields in d.items():
        if rows is None:
            cols = [id_col_name] + list(fields.keys())
            rows = [cols]
        row = [name]
        for k in fields.keys() if rows else []:
            pass
        for c in rows[0][1:]:
            v = fields.get(c, '')
            row.append(json.dumps(v) if isinstance(v, (dict, list)) else v)
        rows.append(row)
    return rows if rows else [[id_col_name]]

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ---- 1. Executive_Result ----
ws = wb.create_sheet('Executive_Result')
exec_rows = [
    ["Field", "Value"],
    ["DER-ID", "PRF-PRIM"],
    ["Closure layer", "C0 -- Primitive/Grammar Reconciliation"],
    ["Question A: mathematically equivalent?", "NO -- no exact equivalence established between any two of the 4 grammars"],
    ["Question B: hierarchically related?", "PARTIAL -- A (DTC) embeds injectively into D (Extended) by direct textual identification (D's own source text: 'same as PRIM-G-00X'); B and C are not shown hierarchically related to A/D or to each other"],
    ["Question C: embeddable in a common grammar?", "PARTIAL -- A embeds in D (see B); B partially/structurally parallels A via an exact external identity (L=Inc*Inc^T) but is not shown embedded; C is not shown embeddable in anything"],
    ["Question D: minimal common representation?", "NOT ESTABLISHED across all 4 grammars. Within grammar A/D alone, {Delta,tau,kappa,Pi} reduces to 3 independently-necessary slots {Delta, tau-slot, kappa}, with Pi derivable and Theta redundant, GIVEN the graph-spectral realization tested here."],
    ["Question E: irreducible primitive(s)?", "Delta -- demonstrated irreducible (cannot be reconstructed from tau+kappa; cospectral non-isomorphic graphs, external theorem). tau-SLOT is irreducible (something transformation-like is required); its FILLER is underdetermined."],
    ["Question F: representational differences?", "Pi vs Theta (in D): shown to coincide with the SAME underlying invariant (ker(L)/connected components) under the graph realization -- likely representational, not substantive, in this realization."],
    ["Question G: physically/structurally meaningful differences?", "Delta's own two source formulations (boundary-operator d^2=0 vs endomorphism required by DER-ORG-002) are TYPE-INCOMPATIBLE -- this is a substantive internal inconsistency, not merely notational."],
    ["Question H: downstream branches depending on reconciliation?", "All of C1-C10 per the corpus's own governance notes cite 'primitive uniqueness remains open' as a caveat on their certified content (see registries/MASTER_CURRENT_CHAT_CANONICAL_RULES.csv). No downstream CERTIFIED result was found to computationally REQUIRE a specific resolution of C0 (the COSMO-DYN chain, e.g., is self-contained given a graph G, independent of how Delta/tau/kappa/Pi are labeled)."],
    ["Question I: does PRF-PRIM close C0?", "NO. PRF-PRIM remains OPEN. This execution narrows the problem (3 grammars -> 4 tested; several sub-claims FALSIFIED; one 3-slot reduction found WITHIN grammar A/D) but does not establish full 4-grammar equivalence, hierarchy, or a common minimal grammar."],
    ["Question J: exact unresolved dependencies remaining?", "See Dependency_Impact sheet. Principal remainder: (1) which realization of tau is intended (automorphism vs. diffusion vs. neither); (2) whether grad(Phi) IS Delta or merely structurally parallel; (3) how PRIM-C (B=(U,V,E)) relates to any other grammar -- ZERO tested or source-documented connection found; (4) Omega vs Psi identity; (5) the type-mismatch in Delta's own definition (boundary op vs endomorphism) is unresolved, not merely observed."],
    ["Overall PRF-PRIM status (this execution)", "PARTIALLY ADVANCED, REMAINS OPEN -- see Status sheet for the full multi-axis classification."],
]
write_rows(ws, exec_rows)

# ---- 2. Primitive_Systems ----
ps = load_json(os.path.join(DOUT, 'primitive_systems.json'))
ws = wb.create_sheet('Primitive_Systems')
rows = [["Grammar_Key", "Label", "Source_ID", "Source_Document", "Composition_Law", "State_Evolution",
          "Existing_Derivation_Status", "Existing_Closure_Status", "Existing_Verification_Status"]]
for gkey, g in ps['grammars'].items():
    rows.append([gkey, g['label'], g['source_id'], g['source_document'], g.get('composition_law', ''),
                 g.get('state_evolution', ''), g['existing_derivation_status'], g['existing_closure_status'],
                 g['existing_verification_status']])
write_rows(ws, rows)
ws2 = wb.create_sheet('_PRF_PRIM_scope')  # small extra note sheet folded at end; keep count controlled below
wb.remove(ws2)

# ---- 3. Definitions ----
ws = wb.create_sheet('Definitions')
def_rows = load_csv(os.path.join(DOUT, 'primitive_systems.csv'))
write_rows(ws, def_rows)

# ---- 4. Type_Signatures ----
ws = wb.create_sheet('Type_Signatures')
write_rows(ws, load_csv(os.path.join(DOUT, 'type_signatures.csv')))

# ---- 5. Compatibility_Matrix ----
ws = wb.create_sheet('Compatibility_Matrix')
write_rows(ws, load_csv(os.path.join(DOUT, 'compatibility_matrix_firstpass.csv')))

# ---- 6. Mapping_Registry ----
ws = wb.create_sheet('Mapping_Registry')
mapping_rows = [
    ["Mapping_ID", "From", "To", "Label", "Status"],
    ["MAP-001", "grad(Phi) [B]", "Delta [A/D] (structural, via Inc^T feeding same L)", "CANDIDATE", "Tested: exact identity L=Inc*Inc^T holds (9/9); identification of grad(Phi) WITH Delta itself NOT shown, only structural parallel"],
    ["MAP-002", "E [B]", "Dirichlet energy functional of grad(Phi) [B]", "CANDIDATE->DEFINITIONAL", "E(Phi)=Phi^T L Phi=||Inc^T Phi||^2, standard identity (ADMITTED EXTERNAL EXT-001)"],
    ["MAP-003", "kappa [A]", "P_ker(L) (spectral projector)", "CANDIDATE, SURVIVES", "Idempotent to machine precision (9/9); commutes with tau1-automorphism (9/9, PROVEN + verified)"],
    ["MAP-004", "Pi [A]", "ker(L)", "CANDIDATE, SURVIVES (derivable)", "= im(kappa) once kappa=P_ker(L); DER-SPC-005 already names R=exp(-beta L) the 'Persistence operator' in source, anticipating this"],
    ["MAP-005", "tau [A] (Liouville-preserving reading)", "graph automorphism", "CANDIDATE, SURVIVES", "Satisfies Liouville condition exactly (det=+-1) on 9/9 families"],
    ["MAP-006", "tau [A] (diffusion reading)", "exp(-tL)", "FALSIFIED", "Violates Liouville condition on 9/9 families (see FALS-001)"],
    ["MAP-007", "Theta [D]", "reachability closure / connected components", "CALCULATED, SURVIVES", "Coincides exactly with dim ker(L) via N-rank(L) on 9/9 families"],
    ["MAP-008", "Omega [D]", "Psi [A] (state variable)", "CANDIDATE, UNRESOLVED", "No computation or recovered source text distinguishes them; not merged by assumption"],
    ["MAP-009", "Delta [A,D] (literal, boundary op. d^2=0)", "simplicial boundary map d1", "CANDIDATE, TYPE-CHECKS FOR ITS OWN DEFINITION", "d1 o d2 = 0 verified exactly on tetrahedron complex; but d1 is NOT an endomorphism of X"],
    ["MAP-010", "Delta [A,D] (functional, DER-ORG-002 usage)", "endomorphism D_hat:X->X, idempotent", "CANDIDATE, TYPE MISMATCH WITH MAP-009", "Required for Psi_(t+1)=kappa(tau(Delta(Psi_t))) to type-check at all"],
    ["MAP-011", "B=(U,V,E) [C]", "(no target found)", "NO CORRESPONDENCE ESTABLISHED", "Zero tested computation or source passage connects PRIM-C-001 to A, B, or D"],
]
write_rows(ws, mapping_rows)

# ---- 7. Mapping_Tests ----
ws = wb.create_sheet('Mapping_Tests')
incid = load_json(os.path.join(DOUT, 'incidence_gradient_results.json'))
rows = [["Graph_family", "n", "m_edges", "L_equals_Inc_IncT_EXACT", "grad_map_injective", "grad_map_surjective_onto_edges", "ker(grad_map)_dim"]]
for name, r in incid.items():
    rows.append([name, r['n'], r['m_edges'], r['L_equals_Inc_IncT'], r['grad_map_injective'], r['grad_map_surjective_onto_edge_space'], r['ker(grad_map)_dim']])
write_rows(ws, rows)

# ---- 8. Composition_Tests ----
ws = wb.create_sheet('Composition_Tests')
ctxt = open(os.path.join(DOUT, 'composition_type_check.txt')).read()
rows = [["Composition_Type_Check_Report (verbatim script output)"]]
for line in ctxt.split('\n'):
    rows.append([line])
write_rows(ws, rows)

# ---- 9. Recursion_Tests ----
ws = wb.create_sheet('Recursion_Tests')
grc = load_json(os.path.join(DOUT, 'grammar_recursion_conjugacy_results.json'))
rows = [["Graph_family", "Delta_alt", "is_zero_map", "fixed_point_dim", "ker_L_dim", "procrustes_residual_vs_exp(-1.0L)"]]
for name, entry in grc.items():
    for label, e in entry.items():
        rows.append([name, label, e['is_zero_map'], e['fixed_point_subspace_dim'], e['ker_L_dim'],
                     round(e['best_orthogonal_procrustes_residual_vs_exp(-1.0*L)'], 4)])
write_rows(ws, rows)

# ---- 10. Invariance_Tests ----
ws = wb.create_sheet('Invariance_Tests')
tau_res = load_json(os.path.join(DOUT, 'tau_liouville_and_commutation_results.json'))
rows = [["Graph_family", "trace_L(2|E|)", "tau2_det_exp(-tL)_at_t=1", "tau2_Liouville_satisfied", "n_automorphisms_checked",
          "tau1_all_dets_+-1", "tau1_Liouville_satisfied", "tau1_kappa_commutation_max_error"]]
for name, r in tau_res.items():
    rows.append([name, r['trace_L_(2|E|)'], r['tau2_det_exp(-tL)_at_various_t']['1.0'], r['tau2_liouville_condition_satisfied'],
                 r['n_automorphisms_checked'], r['tau1_all_dets_are_+-1'], r['tau1_liouville_condition_satisfied'],
                 r['tau1_kappa_commutation_max_error']])
write_rows(ws, rows)

# ---- 11. Minimality_Audit ----
ws = wb.create_sheet('Minimality_Audit')
write_rows(ws, load_csv(os.path.join(DOUT, 'minimality_table.csv')))

# ---- 12. Elimination_Audit ----
ws = wb.create_sheet('Elimination_Audit')
elim = load_json(os.path.join(COUT, 'elimination_audit_results.json'))
rows = [["Test", "Result", "Conclusion"]]
for k, v in elim.items():
    rows.append([k, v['result'], v['conclusion']])
write_rows(ws, rows)

# ---- 13. Common_Grammar ----
ws = wb.create_sheet('Common_Grammar')
gm = load_json(os.path.join(DOUT, 'gamma_min_result.json'))
write_rows(ws, dict_to_rows(gm))

# ---- 14. Graph_Comparison ----
ws = wb.create_sheet('Graph_Comparison')
gg = load_json(os.path.join(GOUT, 'grammar_dependency_graphs.json'))
rows = [["Grammar", "Nodes", "Edges (from -> to : relation)"]]
for gname, gdata in gg.items():
    edges_str = '; '.join(f"{e[0]}->{e[1]}:{e[2]}" for e in gdata['edges'])
    rows.append([gname, ', '.join(gdata['nodes']), edges_str])
write_rows(ws, rows)
mdag = load_json(os.path.join(DOUT, 'master_dag_overlap.json'))
ws.append([])
ws.append(["Master DAG cross-check"])
ws.append(["Primitive Node_IDs confirmed pre-existing in compiler/dag/master_nodes.csv:", ', '.join(mdag['primitive_node_ids_in_master_dag'])])

# ---- 15. Benchmark_Tests ----
ws = wb.create_sheet('Benchmark_Tests')
lap = load_json(os.path.join(COUT, 'laplacian_results.json'))
rows = [["Graph_family", "n", "Spec(L)", "ker_dim", "rank_L", "n_components", "rank=N-c",
          "P_ker_idempotency_error", "limit_exp(-tL)_vs_P_ker_error", "degenerate_spectrum"]]
for name, r in lap.items():
    rows.append([name, r['n_nodes'], str(r['eigenvalues_sorted']), r['ker_dim'], r['rank_L'], r['n_connected_components'],
                 r['rank_L_equals_N_minus_components'], r['P_ker_idempotency_error'],
                 r['lim_exp(-tL)_minus_P_ker_error_at_t=500'], r['degenerate_spectrum']])
write_rows(ws, rows)
K4K33 = open(os.path.join(COUT, 'K4_K33_exact_symbolic.txt')).read()
start = ws.max_row + 3
ws.cell(row=start, column=1, value="Exact symbolic cross-check vs corpus COSMO-BRIDGE-003/004/005:")
for i, line in enumerate(K4K33.split('\n')):
    ws.cell(row=start + 1 + i, column=1, value=line)

# ---- 16. Falsification ----
ws = wb.create_sheet('Falsification')
fals_text = open(os.path.normpath(os.path.join(HERE, '..', '..', '..', 'falsification', 'C0', 'PRF-PRIM', 'FALSIFICATION_RECORDS.md'))).read()
for i, line in enumerate(fals_text.split('\n'), 1):
    ws.cell(row=i, column=1, value=line)
ws.column_dimensions['A'].width = 120

# ---- 17. Existing_Result_Compatibility ----
ws = wb.create_sheet('Existing_Result_Compatibility')
rows = [
    ["Existing corpus result", "Tested against", "Outcome"],
    ["COSMO-BRIDGE-003 (K4: Spec={0,4,4,4}, rank=3)", "Independent sympy exact recomputation", "MATCHES exactly (script 02)"],
    ["COSMO-BRIDGE-004/005 (K3,3: N_bulk=rank(L)=5, d_spectral=5)", "Independent sympy exact recomputation", "MATCHES exactly (script 02)"],
    ["rank(L)=N-c (general claim)", "9 benchmark families incl. disconnected, degenerate-spectrum, torus", "HOLDS on 9/9 (script 02)"],
    ["DER-SPC-005 'Persistence operator R=exp(-beta L)'", "kappa=P_ker(L), Pi=ker(L) candidate mapping", "COMPATIBLE: lim_(t->inf) R(t) = P_ker(L) verified to ~1e-14 on 9/9 families (script 02)"],
    ["MASTER_CURRENT_CHAT_CANONICAL_RULES.csv row C7-001 'Do not assume Fix(Gamma)=Fix(e^{-beta L})'", "tau=exp(-tL) candidate; Gamma_graph vs exp(-hL) conjugacy", "CONFIRMED INDEPENDENTLY: both the Liouville test (FALS-001) and the conjugacy test (FALS-003) support this caution with direct computation"],
    ["BR-001 'Gamma=kappa o tau o Delta, CERTIFIED SPINE'", "Type-check of Delta's dual role (boundary op. vs endomorphism)", "translation fails under construction: literal PRIM-G-001 formula (d1, non-square) cannot compose with kappa,tau (endomorphisms) -- see Composition_Tests. Preserved as CERTIFIED at the LABELING level; flagged as type-incomplete at the operational level."],
    ["DER-ORG-002 'Psi_(t+1)=kappa(tau(Delta(Psi_t))), CERTIFIED'", "Two candidate endomorphism readings of Delta, executed", "translation fails under construction: Delta_alt_1 degenerates the whole composition to the zero map (FALS-002); Delta_alt_2 gives a well-defined but non-conjugate-to-diffusion, fixed-point-free operator (FALS-003). Existing CERTIFIED label NOT erased; operational content flagged OPEN."],
]
write_rows(ws, rows)

# ---- 18. Dependency_Impact ----
ws = wb.create_sheet('Dependency_Impact')
rows = [
    ["Downstream branch", "Computationally requires a specific C0 resolution?", "Basis"],
    ["C1 (fixed-point/organizational closure)", "NOT SHOWN TO REQUIRE", "Fix(Gamma) analysis in the corpus does not depend on which Delta/tau reading is chosen, only on Gamma's abstract composition"],
    ["C2 (spectral/persistence, COSMO-DYN)", "NOT SHOWN TO REQUIRE", "Fully computable from a bare graph G=(V,E) -- confirmed by this DER's own scripts, none of which needed to resolve Delta/tau/kappa's identity to compute Spec(L), ker(L), rank(L)"],
    ["C3 (continuum/Lorentzian)", "NOT SHOWN TO REQUIRE", "Downstream of L alone, same reasoning as C2"],
    ["C4 (constants)", "NOT SHOWN TO REQUIRE", "Constant registry entries reference Seeley-DeWitt/spectral ratios, not primitive-grammar identity"],
    ["C5-C9", "NOT SHOWN TO REQUIRE (this execution did not test these branches directly)", "No computation in this DER touched gauge/matter/quantum/biology/compiler branches; absence of evidence, not evidence of absence"],
    ["C0 itself (this DER)", "Remains OPEN", "See Status sheet"],
    ["New DAG edges actually justified by this execution (see compiler/dag update)", "kappa(P_ker(L)) commutes with tau(automorphism) [PROVEN]; Theta-classes = ker(L)-dim [CALCULATED, 9/9]; L=Inc*Inc^T [PROVEN, external]", "Only these -- no speculative edges added"],
]
write_rows(ws, rows)

# ---- 19. Status ----
ws = wb.create_sheet('Status')
rows = [
    ["Claim", "Derivation_Status", "Universality/Closure_Status"],
    ["A embeds in D (textual)", "ADMITTED (source states 'same as PRIM-G-00X' directly)", "SCOPED to the 4 shared primitives; D's 2 extra primitives (Theta,Omega) not covered"],
    ["kappa=P_ker(L) is idempotent", "CALCULATED (9/9 families, machine precision)", "CALCULATED-UNIVERSAL for symmetric graph Laplacians (spectral theorem guarantees this for ANY such L, not just the 9 tested -- but CANDIDATE that kappa=P_ker(L) is the CORRECT realization of source kappa remains open)"],
    ["kappa(automorphism) commute", "DERIVED (general proof + 9/9 verification)", "UNIVERSAL for this specific pair of realizations (proof is general, not family-limited)"],
    ["tau=exp(-tL) satisfies Liouville", "FALSIFIED (9/9 families)", "NO-GO, general (proof: det=exp(-t*2|E|)->0 for ANY graph with an edge, not just the 9 tested)"],
    ["tau=automorphism satisfies Liouville", "DERIVED (general: permutation matrices are orthogonal, det=+-1)", "UNIVERSAL"],
    ["Pi=ker(L) derivable from kappa", "CALCULATED (9/9)", "CONDITIONAL on kappa=P_ker(L) realization"],
    ["Theta-classes = ker(L)-dim", "CALCULATED (9/9)", "CALCULATED-UNIVERSAL for undirected graphs (reachability=connectivity is a general fact for undirected graphs, not merely observed on 9 examples)"],
    ["L=Inc*Inc^T", "DERIVED/PROVEN (general + 9/9 exact verification)", "UNIVERSAL (standard theorem)"],
    ["Delta reconstructible from tau+kappa", "FALSIFIED (external, admitted)", "NO-GO, general"],
    ["Delta's boundary-op definition vs DER-ORG-002 usage", "CALCULATED (type mismatch demonstrated concretely)", "OPEN -- not resolved which reading is intended by source"],
    ["Gamma_graph conjugate to exp(-hL)", "FALSIFIED (9/9, Procrustes)", "CONDITIONAL NO-GO -- tested only for the specific Delta_alt/tau/kappa realizations and the orthogonal-Procrustes criterion; a different criterion or realization is not ruled out"],
    ["Common minimal grammar Gamma_min across all 4 systems", "OPEN / NOT ESTABLISHED", "OPEN"],
    ["PRF-PRIM overall", "PARTIALLY ADVANCED", "C0 remains OPEN THEOREM (no change to the corpus's own C0 status; this execution adds computed sub-results underneath it)"],
]
write_rows(ws, rows)

# ---- 20. Provenance ----
ws = wb.create_sheet('Provenance')
rows = [
    ["Field", "Value"],
    ["DER-ID", "PRF-PRIM"],
    ["Session", "session_01UBKp9Jq2qXFgravroyUp32"],
    ["Date", "2026-08-17"],
    ["Source records used", "SOURCE-004 (UOC_ToE_Canonical_Bridge_Closure_Execution_v2.0.xlsx): registries/MASTER_PRIMITIVE_REGISTRY.csv, MASTER_PROOF_REGISTRY.csv (PRF-PRIM), 03_RAW_DOC_TABLES.csv, 04_RAW_DOC_PARAGRAPHS.csv (UCG Specification v5.docx Sections 3.1-3.4, 15; Combined Compiler Theories Whitepaper Part IV Section 7.2)"],
    ["External sources used (registered as EXT-*)", "EXT-001: L=Inc*Inc^T (standard algebraic graph theory). EXT-002: cospectral non-isomorphic graphs exist (standard spectral graph theory). EXT-003: automorphism-spectral-projector commutation (standard linear algebra, general argument constructed in this execution)."],
    ["Scripts (reproducibility)", "derivations/C0/PRF-PRIM/01-08_*.py; calculations/C0/PRF-PRIM/01-04_*.py"],
    ["Software", "Python 3, sympy 1.14.0, numpy 2.4.6, networkx 3.6.1"],
    ["Random seed", "N/A -- no randomized algorithm used; all results deterministic (exact sympy or deterministic numpy eigh/SVD)"],
    ["Numerical precision", "float64 (numpy default); exact integer/rational arithmetic used for K4/K3,3 corpus cross-check (sympy)"],
]
write_rows(ws, rows)

# ---- 21. Reproducibility ----
ws = wb.create_sheet('Reproducibility')
rows = [
    ["Step", "Script", "Purpose"],
    ["1", "derivations/C0/PRF-PRIM/01_primitive_systems_recovery.py", "Recover exact primitive definitions from source corpus"],
    ["2", "derivations/C0/PRF-PRIM/02_type_signatures_and_compatibility.py", "Type-theoretic reconciliation, first-pass compatibility matrix"],
    ["3", "calculations/C0/PRF-PRIM/01_graph_families.py", "Build 9-family benchmark graph set"],
    ["4", "calculations/C0/PRF-PRIM/02_laplacian_spectrum_and_persistence.py", "Spec(L), ker(L), rank(L), P_ker(L), corpus cross-check (K4, K3,3)"],
    ["5", "derivations/C0/PRF-PRIM/03_incidence_gradient_identity.py", "L=Inc*Inc^T identity, grad(Phi)<->Delta structural test"],
    ["6", "derivations/C0/PRF-PRIM/04_tau_liouville_and_kappa_commutation.py", "tau Liouville-condition falsification; kappa-tau commutation proof+verification"],
    ["7", "calculations/C0/PRF-PRIM/03_reachability_theta.py", "Theta<->connected components<->ker(L) test"],
    ["8", "calculations/C0/PRF-PRIM/04_elimination_audit.py", "Minimality: is kappa/Pi redundant/derivable?"],
    ["9", "derivations/C0/PRF-PRIM/05_composition_type_check.py", "Delta boundary-operator vs endomorphism type mismatch, tetrahedron complex"],
    ["10", "derivations/C0/PRF-PRIM/06_grammar_recursion_conjugacy_test.py", "Gamma_graph vs exp(-hL) conjugacy test"],
    ["11", "derivations/C0/PRF-PRIM/07_minimal_grammar_and_graph_representation.py", "Minimality table, Gamma_min search, grammar dependency graphs"],
    ["12", "derivations/C0/PRF-PRIM/08_build_workbook.py (this script)", "Assemble this workbook"],
    ["To reproduce", "Run scripts 1-11 in order (each writes its own output/ folder), then run 08_build_workbook.py. All scripts are deterministic; no external network access or observational data used.", ""],
]
write_rows(ws, rows)

os.makedirs(WB_OUT, exist_ok=True)
outfile = os.path.join(WB_OUT, 'C0_PRF_PRIM_Primitive_Reconciliation.xlsx')
wb.save(outfile)
print("Saved", outfile, "with", len(wb.sheetnames), "sheets:")
print(wb.sheetnames)
