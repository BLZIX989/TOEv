import json, csv, os
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = '/home/user/TOEv/'
WB_OUT = REPO + 'results/workbooks/'
os.makedirs(WB_OUT, exist_ok=True)

HEADER_FILL = PatternFill(start_color='7A1F2B', end_color='7A1F2B', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')

def load_json(path):
    with open(path) as f:
        return json.load(f)

def load_csv(path):
    with open(path, newline='') as f:
        return list(csv.reader(f))

def write_rows(ws, rows, start_row=1):
    r = start_row
    if not rows:
        ws.cell(row=r, column=1, value='(no data)')
        return
    header = rows[0]
    for c, val in enumerate(header, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for ri, row in enumerate(rows[1:], start=r + 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=ri, column=c, value=str(val) if not isinstance(val, (int, float, type(None))) else val)
    for c in range(1, len(header) + 1):
        col = get_column_letter(c)
        maxlen = max([len(str(header[c-1]))] + [len(str(row[c-1])) if c-1 < len(row) else 0 for row in rows[1:]])
        ws.column_dimensions[col].width = min(max(12, maxlen * 0.9), 70)
    ws.freeze_panes = 'A2'

def dict_to_rows(d, cols=("Key", "Value")):
    rows = [list(cols)]
    for k, v in d.items():
        rows.append([k, json.dumps(v) if isinstance(v, (dict, list)) else v])
    return rows

def nested_json_to_rows(d, id_col_name="Item"):
    rows = None
    out = []
    for name, fields in d.items():
        if isinstance(fields, dict):
            if rows is None:
                cols = [id_col_name] + list(fields.keys())
                rows = [cols]
            row = [name] + [json.dumps(v) if isinstance(v, (dict, list)) else v for v in fields.values()]
            rows.append(row)
    return rows or [[id_col_name]]

def para_rows(title, paragraphs):
    rows = [[title]]
    for p in paragraphs:
        rows.append([p])
    return rows

# ---- load data ----
morph1 = load_json(REPO + 'calculations/C0_phase2/morphism_test_results.json')
morph2 = load_json(REPO + 'calculations/C0_phase2/morphism_test_results_2.json')
crosswalk = load_csv(REPO + 'results/reconciliation/C0_PRIOR_CALCULATION_CROSSWALK.csv')
obj_reg = load_csv(REPO + 'registries/MASTER_INDEPENDENT_OBJECT_REGISTRY_PHASE2.csv')
dag_edges = load_csv(REPO + 'compiler/dag/MASTER_INDEPENDENT_TOE_DAG_PHASE2.csv')

wb = openpyxl.Workbook()
wb.remove(wb.active)

# 00_README
ws = wb.create_sheet('00_README')
write_rows(ws, para_rows("C0_PRF_PRIM_RECONCILIATION_PHASE2.xlsx", [
    "Phase II, C0/PRF-PRIM reconciliation. Grammars G_A={Delta,tau,kappa,Pi}, "
    "G_B={Delta,tau,kappa}+gradient, G_C={Delta,tau,kappa,Theta,Pi,Omega}.",
    "Governance: never force convergence; every equivalence claim requires explicit mapping+"
    "composition test; CALCULATED != UNIVERSALLY DERIVED; DOWNSTREAM RECOVERY != C0 CLOSURE.",
    "Prior-calculation recovery performed exhaustively across the NX001 family before any new "
    "calculation was designed -- see sheet 11_NX001_CROSSWALK.",
    "Final classification: results/PHASE_2_C0_REPORT.md -- overall C0/PRF-PRIM = C (HIERARCHICAL), "
    "primary evidence G_A<->G_C (F_AC/F_CA, exact, unconditional); G_A<->G_B CONDITIONAL (n=m); "
    "G_B<->G_C OPEN (only trivial composite exists).",
    "Scripts: calculations/C0_phase2/morphism_tests.py (TESTS 1-4), morphism_tests_2.py (TESTS 5-10).",
    "Derivation records: derivations/C0/DER-P2-001_grammar_typing.md, DER-P2-002_morphism_construction.md.",
]))

# 01_STATUS
ws = wb.create_sheet('01_STATUS')
write_rows(ws, [
    ["Grammar_pair", "Relation_type", "DERIVATION_STATUS", "CLOSURE_STATUS", "Scope"],
    ["G_A <-> G_C", "Hierarchical embedding (F_AC, F_CA)", "CALCULATED",
     "HIERARCHICALLY CLOSED -- exact, unconditional, bidirectional (section)", "all 10 test families"],
    ["G_A <-> G_B", "Conditionally equivalent (F_AB, F_BA)", "CALCULATED-CONDITIONAL",
     "CONDITIONALLY CLOSED iff n=m", "3/10 test families satisfy n=m"],
    ["G_B <-> G_C", "No direct translation (F_BC, F_CB fail nontriviality)", "CALCULATED (negative)",
     "OPEN -- minimal obstruction: no non-A-mediated translation constructible", "all 10 test families"],
    ["Gamma_A (literal)", "Composition kappa.tau.Delta", "FALSIFIED", "NO-GO", "all graphs/complexes, n!=m"],
    ["Gamma_A (NX001)", "Composition kappa.tau.Delta, different Delta", "CALCULATED-DERIVED",
     "CALCULATED, scoped", "X={0,1}^(kxk)"],
    ["Gamma_B", "Composition kappa.tau.grad(Phi)", "CALCULATED-CONDITIONAL", "CONDITIONAL", "n=m only"],
    ["Gamma_C", "No formula in corpus", "OPEN", "OPEN", "-"],
    ["Omega vs Psi", "Type identity", "OPEN", "OPEN", "graph + NX001 relational domains"],
])

# 02/03/04 grammars
ws = wb.create_sheet('02_GRAMMAR_A')
write_rows(ws, [r for r in obj_reg if r[2] == 'G_A' or r[0] == 'Object_ID'])
ws = wb.create_sheet('03_GRAMMAR_B')
write_rows(ws, [r for r in obj_reg if r[2] in ('G_B',) or r[0] == 'Object_ID'])
ws = wb.create_sheet('04_GRAMMAR_C')
write_rows(ws, [r for r in obj_reg if r[2] == 'G_C' or r[0] == 'Object_ID'])

# 05_TYPED_OBJECTS
ws = wb.create_sheet('05_TYPED_OBJECTS')
write_rows(ws, obj_reg)

# 06_OPERATOR_SIGNATURES
ws = wb.create_sheet('06_OPERATOR_SIGNATURES')
write_rows(ws, [
    ["Operator", "Domain", "Codomain", "Realization", "Well-typed?"],
    ["Delta (literal)", "C_1 (edges, dim m)", "C_0 (vertices, dim n)", "Inc (boundary map)", "NOT an endomorphism unless n=m"],
    ["Delta (NX001)", "X={0,1}^(kxk)", "X", "elementwise threshold", "YES (endomorphism by construction)"],
    ["tau", "R^n or X", "R^n or X", "automorphism / semigroup / boolean-composition", "YES (endomorphism, 3 realizations)"],
    ["kappa", "R^n or X", "R^n or X", "idempotent projector / closure", "YES (endomorphism)"],
    ["Pi", "-", "subspace of R^n or subset of X", "kernel / fixed-point set", "N/A (a SET, not an operator)"],
    ["nabla / grad(Phi)", "R^n (vertex space)", "R^m (edge space)", "Inc^T", "YES as an operator; composes with kappa/tau only if n=m"],
    ["Theta", "vertex set or X", "partition of domain", "reachability closure / basins", "YES (endomorphism-free, well-defined partition map)"],
    ["Omega", "-", "same space as Psi", "state-space element", "type-indistinguishable from Psi"],
])

# 07_MORPHISMS
ws = wb.create_sheet('07_MORPHISMS')
write_rows(ws, [
    ["Translation", "Domain", "Codomain", "Action", "Well-defined?", "Preservation_verdict"],
    ["F_AB", "realizations of Delta", "realizations of grad(Phi)", "matrix transpose Inc->Inc^T", "YES, unconditional",
     "Rank+nonzero spectrum: PASS unconditional (10/10 + tetrahedron d2). Kernel/zero-block: conditional n=m."],
    ["F_BA", "realizations of grad(Phi)", "realizations of Delta", "matrix transpose Inc^T->Inc", "YES, unconditional",
     "Involution/identity law: PASS, trivial (10/10)."],
    ["F_AC", "ker(L) (Pi)", "Theta-partition", "cardinality map", "YES, unconditional", "Cardinality-preserving PASS (10/10)."],
    ["F_CA", "Theta-partition", "ker(L) (Pi)", "span of component indicator vectors", "YES, unconditional",
     "EXACT canonical reconstruction PASS (10/10) -- strongest positive result."],
    ["F_BC", "realizations of grad(Phi)", "Theta-partition", "composite F_AC.F_BA only", "Only as composite",
     "FAILS nontriviality test -- SCOPE-DEPENDENT, only indirect/trivial."],
    ["F_CB", "Theta-partition, Omega", "realizations of grad(Phi)", "composite F_AB.F_CA only", "Only as composite",
     "FAILS nontriviality test -- SCOPE-DEPENDENT, only indirect/trivial."],
])

# 08_COMPOSITION
ws = wb.create_sheet('08_COMPOSITION')
write_rows(ws, [
    ["Composition tested", "Result", "Families tested", "n passing", "Note"],
    ["F_BA o F_AB = id (involution)", "PASS, unconditional", "10", "10/10", "Trivial linear-algebra fact (transpose is an involution)."],
    ["F_AC o F_CA = id (exact section)", "PASS, unconditional", "10", "10/10", "Nontrivial: recovers ORIGINAL partition exactly, not just cardinality."],
    ["F_BC := F_AC . F_BA", "Well-defined only as composite; carries no grad(Phi)-specific info", "10", "10/10 (all trivial)", "FAILS nontriviality -- see TEST 9."],
    ["F_CB := F_AB . F_CA", "Well-defined only as composite; carries no Theta/Omega-specific info", "10", "10/10 (all trivial)", "FAILS nontriviality -- see TEST 9."],
])

# 09_FIXED_POINTS
ws = wb.create_sheet('09_FIXED_POINTS')
fp = morph2['TEST_10_fixed_point_preservation_Gamma_A_vs_Gamma_B']
rows = [["Family", "n", "m", "dim_Fix(Gamma_A)", "dim_Fix(Gamma_B)", "Equal?"]]
for name, r in fp['per_family'].items():
    rows.append([name, r['n'], r['m'], r['dim_Fix(Gamma_A)_(ker_L_vertex)'], r['dim_Fix(Gamma_B)_(ker_L_edge)'], r['fixed_point_dims_equal']])
rows.append([])
rows.append([f"Matched {fp['n_families_matched']}/{fp['n_families_typechecking_(n=m)']} n=m-typechecking families. "
             f"NX001's own Fix(Gamma)=5 (Bell number B_3) result cited, not recomputed."])
write_rows(ws, rows)

# 10_SPECTRAL_PRESERVATION
ws = wb.create_sheet('10_SPECTRAL_PRESERVATION')
sp = morph2['TEST_5_spectral_preservation_F_AB']
rows = [["Family", "n", "m", "Nonzero_spectrum_preserved", "Zero_mult_vertex", "Zero_mult_edge", "Predicted_diff_(m-n)", "Prediction_matches"]]
for name, r in sp['per_family'].items():
    rows.append([name, r['n'], r['m'], r['nonzero_spectrum_preserved_by_F_AB'],
                 r['zero_eigenvalue_multiplicity_L_vertex_(dim_ker_Delta^T)'],
                 r['zero_eigenvalue_multiplicity_L_edge_(dim_ker_Delta)'],
                 r['predicted_difference_(m-n)'], r['prediction_matches']])
rows.append([])
rows.append(["Tetrahedron d2-level (face/edge) test:", json.dumps(morph2['TEST_5b_tetrahedron_2complex_d2_level']['nonzero_spectrum_preserved'])])
rows.append(["Heat semigroup preservation (t=0.37), K4 and K3_NEW:", json.dumps(morph2['TEST_6_heat_semigroup_preservation'])])
write_rows(ws, rows)

# 11_NX001_CROSSWALK
ws = wb.create_sheet('11_NX001_CROSSWALK')
write_rows(ws, crosswalk)

# 12_COUNTEREXAMPLES
ws = wb.create_sheet('12_COUNTEREXAMPLES')
write_rows(ws, [
    ["Structure", "Used for", "Result"],
    ["K4", "F_AB/Gamma_B typing, spectral preservation", "n=4,m=6, type mismatch; nonzero spectrum preserved"],
    ["K3,3", "F_AB/Gamma_B typing, spectral preservation", "n=6,m=9, type mismatch; nonzero spectrum preserved"],
    ["C6 (6-cycle)", "F_AB/Gamma_B typing", "n=6,m=6 -- Gamma_B TYPE-CHECKS (rare n=m case)"],
    ["P5 (path)", "F_AB/Gamma_B typing", "n=5,m=4, type mismatch"],
    ["disjoint K4 (sqcup) K3,3", "F_AB/Gamma_B typing", "type mismatch"],
    ["two disjoint triangles", "F_AB/Gamma_B typing", "n=6,m=6 -- Gamma_B TYPE-CHECKS"],
    ["torus 4x4 lattice", "F_AB/Gamma_B typing, periodic lattice requirement", "type mismatch"],
    ["Petersen graph", "F_AB/Gamma_B typing", "type mismatch"],
    ["star K1,5", "F_AB/Gamma_B typing", "type mismatch"],
    ["K3 (complete graph on 3, newly added)", "F_AB/Gamma_B typing", "n=3,m=3 -- Gamma_B TYPE-CHECKS"],
    ["Tetrahedron 2-complex (d1, d2)", "Composition type check (PRF-PRIM); F_AB spectral preservation (Phase II, d2 level)", "d1.d2=0 exact; nonzero spectrum preserved at d2 level -- non-1-skeleton structure"],
    ["NX001 relational system X={0,1}^(3x3)", "Non-graph organizational structure (required by directive)", "grad(Phi) NOT instantiable; Theta partially instantiable (basins); Omega directly instantiable (=state)"],
])

# 13_PROOFS
ws = wb.create_sheet('13_PROOFS')
write_rows(ws, [
    ["Proof", "Statement", "Method", "Location"],
    ["F_CA exact reconstruction", "Connected-component indicator vectors form a canonical basis of ker(L); "
     "reading off supports recovers the original Theta-partition exactly", "Direct construction + verification, 10/10 families",
     "calculations/C0_phase2/morphism_tests_2.py TEST 7"],
    ["F_AB nonzero-spectrum transport", "Nonzero eigenvalues of M.M^T and M^T.M coincide with multiplicity for any real matrix M; "
     "zero-eigenvalue multiplicities differ by exactly |n-m|", "Classical linear algebra fact, verified numerically, "
     "10 graphs + tetrahedron d2 level", "calculations/C0_phase2/morphism_tests_2.py TEST 5, 5b"],
    ["F_BA involution", "(Inc^T)^T = Inc exactly", "Trivial, verified 10/10", "calculations/C0_phase2/morphism_tests_2.py TEST 8"],
    ["NX001E general closure theorem (cited, not rederived)", "Gamma_A*(A)=Gamma_B*(A)=Eq(A) for all finite V",
     "Closure-operator argument (extensivity+monotonicity+finite-chain termination); independently logically verified sound this phase",
     "source_records/spreadsheets/NX001_family/ (NX001E); results/reconciliation/C0_PRIOR_CALCULATION_CROSSWALK.csv row NX001E-THEOREM"],
])

# 14_FALSIFICATIONS
ws = wb.create_sheet('14_FALSIFICATIONS')
write_rows(ws, [
    ["Claim falsified", "Falsification method", "Location"],
    ["Gamma_A=kappa.tau.Delta well-typed (literal Delta reading)", "Domain(6)!=Codomain(4) demonstrated on tetrahedron complex",
     "derivations/C0/PRF-PRIM/05_composition_type_check.py (prior, cited)"],
    ["Gamma_B=kappa.tau.grad(Phi) well-typed in general", "7/10 test families have n!=m, demonstrated concretely",
     "calculations/C0_phase2/morphism_tests.py TEST 2"],
    ["F_BC, F_CB are independent (non-composite) morphisms", "No construction found that does not factor through A; "
     "explicit nontriviality test fails on 10/10 families", "calculations/C0_phase2/morphism_tests_2.py TEST 9"],
    ["NX001F's broader spectral-invariance conjecture F7 (cited, not rederived)", "NX001G's own falsification: clique-graph vs "
     "incidence-graph functors give different spectra for the same equivalence-relation semantics",
     "results/reconciliation/C0_PRIOR_CALCULATION_CROSSWALK.csv row NX001G-FALSIFICATION"],
])

# 15_OBSTRUCTIONS
ws = wb.create_sheet('15_OBSTRUCTIONS')
write_rows(ws, [
    ["Obstruction", "Minimal_formal_statement", "Blocks"],
    ["Gamma_B typing requires n=m", "kappa,tau act on vertex space (dim n); grad(Phi) codomain is edge space (dim m); "
     "composition kappa.tau.grad(Phi) type-checks iff n=m", "Full G_A<->G_B equivalence (only conditional)"],
    ["No direct B<->C translation", "Every attempted F_BC/F_CB construction factors through A and discards all "
     "grad(Phi)/Theta/Omega-specific content in the process", "Full three-grammar (G_A,G_B,G_C) simultaneous closure"],
    ["Gamma_C undefined", "No composed formula for kappa.tau.Delta.Theta... (or any subset) appears anywhere in the corpus",
     "Any claim about Gamma_C's well-typedness, fixed points, or spectral behavior"],
    ["Omega vs Psi type-identity", "In every domain tested (graph, NX001 relational), Omega's stated definition "
     "coincides exactly with Psi's; no test constructed so far distinguishes them as different TYPES",
     "Whether G_C's primitive set is genuinely larger than G_A's, or only nominally so"],
])

# 16_DOWNSTREAM_IMPACT
ws = wb.create_sheet('16_DOWNSTREAM_IMPACT')
write_rows(ws, [
    ["Downstream obligation", "Depends on", "Status", "Recommendation"],
    ["C1 / DER-ORG-006 (fixed-point equivalence theorem)", "C0 PRF-PRIM classification", "C0 = C (HIERARCHICAL), primary path CLOSED; "
     "G_B and G_C sub-paths CONDITIONAL/OPEN", "Do NOT open C1 broadly; MAY use the G_A<->G_C closed result "
     "as a scoped prerequisite if a future C1 obligation only requires that pair"],
    ["Any future claim of full three-grammar universality", "Resolution of G_B<->G_C obstruction and Gamma_C's definition",
     "OPEN", "Next executable calculation identified in PHASE_2_C0_REPORT.md; NOT executed this phase"],
])

# 17_PROVENANCE
ws = wb.create_sheet('17_PROVENANCE')
write_rows(ws, [
    ["Source", "Role", "SOURCE_ID"],
    ["UOC_ToE_NX001_Calculation_Package.xlsx", "prior calculation recovery", "SOURCE-011"],
    ["UOC_ToE_NX001B_Extensional_Equivalence_Calculation.xlsx", "prior calculation recovery", "SOURCE-012"],
    ["UOC_ToE_NX001E_Formal_Normal_Form_Equivalence.xlsx", "prior calculation recovery (cited theorem)", "SOURCE-013"],
    ["UOC_ToE_NX001F_Spectral_Descent_Equivalence.xlsx", "prior calculation recovery (cited theorems T1-T7)", "SOURCE-014"],
    ["UOC_ToE_NX001G_Independent_Graph_Extraction_Spectral_Invariance.xlsx", "prior calculation recovery (falsification)", "SOURCE-015"],
    ["derivations/C0/PRF-PRIM/ (all scripts)", "Phase-I C0 baseline, cited not recomputed", "-"],
    ["calculations/C0_phase2/morphism_tests.py, morphism_tests_2.py", "New Phase-II computation", "PHASE2-C0-EXEC-2026-08-17"],
    ["derivations/C0/DER-P2-001_grammar_typing.md", "New Phase-II derivation record", "DER-P2-001"],
    ["derivations/C0/DER-P2-002_morphism_construction.md", "New Phase-II derivation record", "DER-P2-002"],
])

wb.save(WB_OUT + 'C0_PRF_PRIM_RECONCILIATION_PHASE2.xlsx')
print("Saved with sheets:", wb.sheetnames)
print("n sheets:", len(wb.sheetnames))
